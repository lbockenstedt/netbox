"""Pure-parse Excel rack-layout detection for the NetBox importer.

No NetBox calls live here — this module turns an .xlsx workbook into a JSON-safe
description of the rack sheets it contains (so the WebUI can render a
column-mapping UI) and, later, turns one sheet + a user column-map into device
rows for ``DcimMixin.import_rack_layout``.

Two sheet shapes are recognized (the HPE DXP lab workbook uses both):

* **one-rack-per-sheet** — a header row (``RU``, ``F/R``, ``Type of device``,
  ``Hostname``, …) followed by one row per device slot. This is the clean,
  high-value shape; the user maps columns → NetBox fields.
* **summary** — a whole-lab elevation grid (``MIP Rack Layout``) where
  rack-blocks are stacked vertically, each introduced by a ``RACK <name>``
  cell with ``RU`` / ``Front`` / ``Rear`` columns. Front/Rear text cells become
  front/rear-face devices named by their text (no model/serial/IP).

openpyxl is imported lazily so a spoke without the dep degrades to a clear
ERROR rather than crashing on import.
"""
from __future__ import annotations

import logging
import re
from typing import Any, Dict, List, Optional

logger = logging.getLogger("NetboxEngine")

# ─── column-label → importer-field synonyms ─────────────────────────────────
# Each field lists header labels (lowercased, whitespace-collapsed) that map to
# it. The first matching label wins. Order matters: more-specific fields first.
_FIELD_SYNONYMS: Dict[str, List[str]] = {
    "position":    ["ru", "rack unit", "u", "rack u"],
    "face":        ["f/r", "fr", "front/rear", "f/r "],
    "device_type": ["type of device", "device type", "model", "type"],
    "name":        ["hostname", "name", "host name", "device name"],
    "serial":      ["serial number", "serial", "serial no", "s/n"],
    "asset_tag":   ["part number", "asset tag", "part no", "p/n"],
    "mac":         ["mac address", "mac", "mac addr"],
    "mgmt_ip":     ["mgmt ip", "management ip", "mgmt ip address", "local ip",
                    "ip address", "ip"],
    "role":        ["function", "role"],
    "status":      ["status"],
    "description": ["description", "desc", "comments"],
}

# The fixed set of target fields the WebUI offers in its mapping <select>.
TARGET_FIELDS: List[str] = (
    ["name", "device_type", "serial", "asset_tag", "position", "face",
     "mgmt_ip", "mac", "role", "status", "description", "ignore"]
)

_PREVIEW_ROWS = 20  # cap on preview rows returned to the WebUI per sheet


def _norm_label(v: Any) -> str:
    """Lowercase + collapse whitespace + strip, for header matching."""
    return re.sub(r"\s+", " ", str(v or "").strip().lower())


def _cell_str(v: Any) -> str:
    """Stringify a cell value, stripping whitespace; '' for None/blank."""
    if v is None:
        return ""
    s = str(v).strip()
    # openpyxl can yield trailing '\t' / nbsp in messy sheets; normalize.
    s = s.replace("\t", " ").replace(" ", " ")
    return re.sub(r"\s+", " ", s).strip()


def _guess_field(label: str) -> Optional[str]:
    """Return the importer field a header label maps to, or None."""
    n = _norm_label(label)
    if not n:
        return None
    for field, syns in _FIELD_SYNONYMS.items():
        if n in syns:
            return field
    # substring fallback: e.g. "MGMT IP (OOB)" contains "mgmt ip"
    for field, syns in _FIELD_SYNONYMS.items():
        for s in syns:
            if s and s in n:
                return field
    return None


def load_workbook_from_bytes(data: bytes):
    """Parse .xlsx bytes into an openpyxl Workbook (lazy import)."""
    import io
    try:
        import openpyxl
    except ImportError as e:  # pragma: no cover - env-dependent
        raise RuntimeError(
            "openpyxl is required to parse .xlsx imports; install it on the "
            "netbox spoke (`pip install openpyxl`) and restart.") from e
    return openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)


def _is_header_row(cells: List[str]) -> bool:
    """A one-rack-per-sheet header row contains 'RU' AND a device/hostname col."""
    labels = {_norm_label(c) for c in cells}
    if "ru" not in labels:
        return False
    return any(k in labels for k in
               ("hostname", "type of device", "device type", "model", "name"))


def _uheight_from_metadata(rows: List[List[Any]]) -> Optional[int]:
    """Scan for a trailing rack-metadata row like '4-Post 51U Rack' → 51."""
    for cells in rows[-12:]:
        for c in cells:
            s = _cell_str(c)
            m = re.search(r"(\d+)\s*u\b", s, re.IGNORECASE)
            if m and "rack" in s.lower():
                try:
                    return int(m.group(1))
                except ValueError:
                    continue
    return None


def _detect_one_rack_sheet(ws) -> Optional[Dict[str, Any]]:
    """Detect a one-rack-per-sheet shape. Returns a rack_sheet dict or None."""
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    header_idx = None
    for i, cells in enumerate(rows):
        if _is_header_row([_cell_str(c) for c in cells]):
            header_idx = i
            break
    if header_idx is None:
        return None
    header = [(_cell_str(c) if j < len(rows[header_idx]) else "")
              for j, c in enumerate(rows[header_idx])]
    # Build label → column index (first occurrence wins).
    label_to_idx: Dict[str, int] = {}
    for idx, label in enumerate(header):
        lbl = _cell_str(label)
        if lbl and lbl not in label_to_idx:
            label_to_idx[lbl] = idx
    # Guess a column_map: source-label → target-field.
    column_map: Dict[str, str] = {}
    for label, idx in label_to_idx.items():
        f = _guess_field(label)
        if f:
            column_map[label] = f
    # Resolve the RU column index for U-height + device-row detection.
    ru_idx = None
    for label, idx in label_to_idx.items():
        if _guess_field(label) == "position":
            ru_idx = idx
            break
    # Device rows: below header, with a numeric RU and at least one non-empty
    # device-ish cell (hostname / type / function). Nameless rows are still
    # surfaced (the importer skips rows lacking name+serial, but the user sees
    # them in the preview so they know to fill names or map a column).
    device_rows: List[List[Any]] = []
    max_ru = 0
    name_idx = next((i for lbl, i in label_to_idx.items()
                     if _guess_field(lbl) == "name"), None)
    type_idx = next((i for lbl, i in label_to_idx.items()
                     if _guess_field(lbl) == "device_type"), None)
    role_idx = next((i for lbl, i in label_to_idx.items()
                     if _guess_field(lbl) == "role"), None)
    for cells in rows[header_idx + 1:]:
        ru_val = cells[ru_idx] if ru_idx is not None and ru_idx < len(cells) else None
        try:
            ru = int(ru_val) if ru_val is not None and str(ru_val).strip() != "" else None
        except (ValueError, TypeError):
            ru = None
        if ru is not None and ru > max_ru:
            max_ru = ru
        # stop scanning at the trailing metadata row (RU == 'N/A')
        if _norm_label(ru_val) in ("n/a", "na", ""):
            # could be a blank separator or the metadata row; skip but continue
            continue
        has_device = any(
            _cell_str(cells[i]) for i in (name_idx, type_idx, role_idx)
            if i is not None and i < len(cells))
        if ru is not None and has_device:
            device_rows.append([_cell_str(c) for c in cells])
    # Rack U-height = the rack's total U count. The trailing metadata row
    # ('4-Post 51U Rack') gives the true size; the max occupied RU is a lower
    # bound. Take the larger so a sparsely-filled 51U rack still reports 51.
    meta = _uheight_from_metadata(rows) or 0
    u_height = max(max_ru, meta) or max_ru or 0
    return {
        "shape": "one-rack",
        "sheet": ws.title,
        "rack_name": ws.title,
        "u_height": u_height,
        "columns": [h for h in header if h],
        "column_map": column_map,
        "preview_rows": device_rows[:_PREVIEW_ROWS],
        "device_count": len(device_rows),
    }


_RACK_BLOCK_RE = re.compile(r"^\s*rack\s+(.+)$", re.IGNORECASE)


def _detect_summary_sheets(ws) -> List[Dict[str, Any]]:
    """Detect rack-blocks in a whole-lab summary sheet. Returns a list of
    rack_sheet dicts (one per block). Best-effort: handles vertically-stacked
    blocks (``RACK <name>`` … RU/Front/Rear … rows …). Side-by-side rack groups
    in one row band are not handled in v1 — the user can deselect these."""
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    sheets: List[Dict[str, Any]] = []
    n = len(rows)
    for i, cells in enumerate(rows):
        for c in cells:
            m = _RACK_BLOCK_RE.match(_cell_str(c))
            if not m:
                continue
            rack_name = m.group(1).strip()
            # nearest 'RU' header at/after this row
            block = _parse_summary_block(rows, i, rack_name, ws.title)
            if block:
                sheets.append(block)
            break
    return sheets


def _parse_summary_block(rows: List[List[Any]], start: int,
                         rack_name: str, sheet: str) -> Optional[Dict[str, Any]]:
    """From a 'RACK <name>' row, find the RU/Front/Rear columns and read device
    rows until the next rack block or a long blank band."""
    n = len(rows)
    # find the RU header row within the next ~6 rows
    ru_col = front_col = rear_col = None
    header_row = None
    for j in range(start, min(start + 8, n)):
        cells = rows[j]
        labels = [(_norm_label(c), idx) for idx, c in enumerate(cells)]
        ru = next((idx for lbl, idx in labels if lbl == "ru"), None)
        if ru is not None:
            ru_col = ru
            header_row = j
            # Front/Rear headers usually sit in the row after RU, same/next cols.
            # Look in this row and the next for 'front'/'rear'.
            for k in (j, j + 1):
                if k >= n:
                    break
                klabels = [(_norm_label(c), idx) for idx, c in enumerate(rows[k])]
                if front_col is None:
                    front_col = next((idx for lbl, idx in klabels if lbl == "front"), None)
                if rear_col is None:
                    rear_col = next((idx for lbl, idx in klabels if lbl == "rear"), None)
            break
    if ru_col is None:
        return None
    if front_col is None:
        front_col = ru_col + 1
    if rear_col is None:
        rear_col = ru_col + 2
    device_rows: List[List[Any]] = []
    max_ru = 0
    blank_run = 0
    for cells in rows[header_row + 1:]:
        ru_val = cells[ru_col] if ru_col < len(cells) else None
        try:
            ru = int(ru_val) if ru_val is not None and str(ru_val).strip() != "" else None
        except (ValueError, TypeError):
            ru = None
        # Stop at the next rack block.
        if any(_RACK_BLOCK_RE.match(_cell_str(c)) for c in cells):
            break
        if ru is None and not any(_cell_str(c) for c in cells):
            blank_run += 1
            if blank_run >= 4:
                break
            continue
        blank_run = 0
        if ru is None:
            continue
        if ru > max_ru:
            max_ru = ru
        front = _cell_str(cells[front_col]) if front_col < len(cells) else ""
        rear = _cell_str(cells[rear_col]) if rear_col < len(cells) else ""
        if front or rear:
            device_rows.append([str(ru), front, rear])
    if not device_rows:
        return None
    # device_count = front+rear device slots (a RU with both front and rear
    # text yields two devices), not raw rows.
    n_devices = sum(1 for r in device_rows if r[1]) + sum(1 for r in device_rows if r[2])
    return {
        "shape": "summary",
        "sheet": sheet,
        "rack_name": rack_name,
        "u_height": max_ru or 0,
        "columns": ["RU", "Front", "Rear"],
        "column_map": {"Front": "name", "Rear": "name"},  # informational
        "preview_rows": device_rows[:_PREVIEW_ROWS],
        "device_count": n_devices,
    }


def detect_rack_sheets(wb) -> List[Dict[str, Any]]:
    """Scan every worksheet and return detected rack-sheet descriptors.

    A sheet may yield zero, one (one-rack-per-sheet), or many (summary blocks)
    descriptors. Order preserves the workbook's sheet order."""
    out: List[Dict[str, Any]] = []
    for ws in wb.worksheets:
        title = (ws.title or "").strip()
        if not title or title.lower() in ("toc", "calcs"):
            continue
        # Try one-rack-per-sheet first; if not, try summary blocks.
        one = _detect_one_rack_sheet(ws)
        if one:
            out.append(one)
            continue
        out.extend(_detect_summary_sheets(ws))
    return out


# ─── commit-time parse: one sheet + column_map → device field dicts ─────────

def parse_one_rack_sheet(ws, column_map: Dict[str, str]) -> Dict[str, Any]:
    """Re-read a one-rack-per-sheet worksheet and apply the user's column_map
    (source-label → target-field) to produce per-device field dicts.

    Returns ``{rack_name, u_height, devices: [{field: value, ...}]}``. Rows
    with no name AND no serial are omitted (returned as ``skipped`` count)."""
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    header_idx = None
    for i, cells in enumerate(rows):
        if _is_header_row([_cell_str(c) for c in cells]):
            header_idx = i
            break
    if header_idx is None:
        return {"rack_name": ws.title, "u_height": 0, "devices": [], "skipped": 0,
                "error": "header row not found on re-parse"}
    header = [_cell_str(c) for c in rows[header_idx]]
    label_to_idx: Dict[str, int] = {}
    for idx, label in enumerate(header):
        if label and label not in label_to_idx:
            label_to_idx[label] = idx
    # Invert the user map → field → column index (last label wins per field).
    field_to_idx: Dict[str, int] = {}
    for src_label, field in (column_map or {}).items():
        if field in (None, "", "ignore"):
            continue
        idx = label_to_idx.get(src_label)
        if idx is None:
            # try a normalized match (label may have been trimmed in the UI)
            idx = next((i for lbl, i in label_to_idx.items()
                        if _norm_label(lbl) == _norm_label(src_label)), None)
        if idx is not None:
            field_to_idx[field] = idx
    ru_idx = field_to_idx.get("position")
    devices: List[Dict[str, Any]] = []
    skipped = 0
    max_ru = 0
    for cells in rows[header_idx + 1:]:
        ru_val = cells[ru_idx] if ru_idx is not None and ru_idx < len(cells) else None
        try:
            ru = int(ru_val) if ru_val is not None and str(ru_val).strip() != "" else None
        except (ValueError, TypeError):
            ru = None
        if _norm_label(ru_val) in ("n/a", "na"):
            continue
        if ru is not None and ru > max_ru:
            max_ru = ru
        dev: Dict[str, Any] = {}
        for field, idx in field_to_idx.items():
            if idx is not None and idx < len(cells):
                v = _cell_str(cells[idx])
                if v:
                    dev[field] = v
        # Require a RU and at least a name or serial to import a device.
        if ru is None or not (dev.get("name") or dev.get("serial")):
            if ru is not None and (dev.get("device_type") or dev.get("role")):
                skipped += 1  # an occupied slot we can't name — surfaced
            continue
        dev["position"] = ru
        devices.append(dev)
    return {"rack_name": ws.title, "u_height": max_ru or 0,
            "devices": devices, "skipped": skipped}


def parse_summary_block_by_name(wb, rack_name: str) -> Optional[Dict[str, Any]]:
    """Re-read a summary sheet and extract the front/rear device rows for the
    named rack block. Returns ``{rack_name, u_height, devices}`` or None."""
    for ws in wb.worksheets:
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        for i, cells in enumerate(rows):
            for c in cells:
                m = _RACK_BLOCK_RE.match(_cell_str(c))
                if m and m.group(1).strip() == rack_name:
                    block = _parse_summary_block(rows, i, rack_name, ws.title)
                    if not block:
                        return None
                    return _summary_rows_to_devices(block, rack_name)
    return None


def _summary_rows_to_devices(block: Dict[str, Any], rack_name: str) -> Dict[str, Any]:
    """Turn a summary block's [RU, Front, Rear] rows into front+rear devices."""
    devices: List[Dict[str, Any]] = []
    for row in block["preview_rows"]:
        ru_s, front, rear = (row + ["", "", ""])[:3]
        try:
            ru = int(ru_s)
        except (ValueError, TypeError):
            continue
        if front:
            devices.append({"name": front, "position": ru, "face": "front"})
        if rear:
            devices.append({"name": rear, "position": ru, "face": "rear"})
    return {"rack_name": rack_name, "u_height": block.get("u_height", 0),
            "devices": devices, "skipped": 0}